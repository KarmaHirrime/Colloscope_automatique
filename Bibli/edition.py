import xml.etree.ElementTree as ET
from classesColloscopeAuto import *
from equations import *
import os
from fpdf import FPDF
from collections import defaultdict
from datetime import timedelta
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path





def minutes_to_hhmm(minutes):
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"






def export_resultats_pdf(chemin_export, resultats, db):
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    import os

    def init_unicode_pdf(font_size=11):
        font_path = os.path.join("fonts", "DejaVuSans.ttf")
        if not os.path.exists(font_path):
            raise FileNotFoundError("Le fichier DejaVuSans.ttf est manquant dans le dossier 'fonts'.")
        pdf = FPDF()
        pdf.add_font("DejaVu", "", font_path)
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(left=15, top=15, right=15)
        pdf.add_page()
        pdf.set_font("DejaVu", "", font_size)
        return pdf

    def minutes_to_hhmm(minutes):
        h, m = divmod(minutes, 60)
        return f"{h:02d}:{m:02d}"

    def render_activity_line(pdf, texte):
        pdf.set_x(pdf.l_margin + 2)
        pdf.multi_cell(0, 8, f"– {texte}")

    def render_eleve_line(pdf, texte):
        pdf.set_x(pdf.l_margin + 4)
        pdf.multi_cell(0, 8, f"  {texte}")

    os.makedirs(os.path.join(chemin_export, "colleurs"), exist_ok=True)
    for classe in db["classes"].values():
        os.makedirs(os.path.join(chemin_export, classe.nom), exist_ok=True)

    calendrier = db["calendrier"]
    TP_class = db["TPs"]

    for etudiant in db["etudiants"].values():
        pdf = init_unicode_pdf(font_size=11)
        pdf.set_font("DejaVu", "", 14)
        pdf.cell(0, 10, f"Planning de {etudiant.nom}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.ln(5)
        pdf.set_font("DejaVu", "", 11)

        for semaine in sorted(resultats.keys()):
            activites = resultats[semaine].get(etudiant, [])
            if not activites:
                continue

            tri = sorted(activites, key=lambda t: (t[2].jour, t[2].debut))

            pdf.set_font("DejaVu", "", 12)
            pdf.cell(0, 8, f"— Semaine {semaine} —", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
            pdf.set_draw_color(180)
            x1, y = pdf.get_x(), pdf.get_y()
            pdf.line(x1, y, 200, y)
            pdf.ln(2)
            pdf.set_font("DejaVu", "", 11)

            for activite, colleur, horaire in tri:
                jour = horaire.jour
                date = calendrier.get_date(semaine, jour).strftime('%d/%m/%Y')
                debut = minutes_to_hhmm(horaire.debut)
                fin = minutes_to_hhmm(horaire.debut + horaire.duree)
                salle=horaire.salle

                if isinstance(activite, TP):
                    texte = f"{jour} {date}, {debut}–{fin} : {activite.nom}"
                else:
                    texte = f"{jour} {date}, {debut}–{fin}"
                    if activite.dureePreparation:
                        passage = minutes_to_hhmm(horaire.debut + activite.dureePreparation)
                        texte += f" (passage à {passage})"
                    if salle=="":
                        texte += f" : {activite.nom} avec {colleur.nom}"
                    else:
                        texte += f" : {activite.nom} avec {colleur.nom} en salle {salle}"

                render_activity_line(pdf, texte)

            pdf.ln(5)

        filepath = os.path.join(chemin_export, etudiant.classe.nom, f"{etudiant.nom}.pdf")
        pdf.output(filepath)

    planning_colleurs = {}
    for semaine, dico_etu in resultats.items():
        for etu, activites in dico_etu.items():
            for act, colleur, horaire in activites:
                if colleur is None:
                    continue
                key = (colleur, semaine)
                planning_colleurs.setdefault(key, []).append((horaire, act, etu))

    colleurs_dict = {}
    for (colleur, semaine), lignes in planning_colleurs.items():
        colleurs_dict.setdefault(colleur, {})[semaine] = lignes

    for colleur, semaines_dict in colleurs_dict.items():
        pdf = init_unicode_pdf(font_size=11)
        pdf.set_font("DejaVu", "", 14)
        pdf.cell(0, 10, f"Planning de {colleur.nom}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.ln(5)
        pdf.set_font("DejaVu", "", 11)

        for semaine in sorted(semaines_dict.keys()):
            lignes = semaines_dict[semaine]
            horaires_groupes = {}

            for horaire, matiere, etu in lignes:
                horaires_groupes.setdefault(horaire, []).append((matiere, etu))

            if not horaires_groupes:
                continue

            pdf.set_font("DejaVu", "", 12)
            pdf.cell(0, 8, f"— Semaine {semaine} —", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
            pdf.set_draw_color(180)
            x1, y = pdf.get_x(), pdf.get_y()
            pdf.line(x1, y, 200, y)
            pdf.ln(2)
            pdf.set_font("DejaVu", "", 11)

            for horaire in sorted(horaires_groupes, key=lambda h: (h.jour, h.debut)):
                jour = horaire.jour
                date = calendrier.get_date(semaine, jour).strftime('%d/%m/%Y')
                debut = minutes_to_hhmm(horaire.debut)
                fin = minutes_to_hhmm(horaire.debut + horaire.duree)
                matiere = horaires_groupes[horaire][0][0]
                salle=horaire.salle
                if salle=="":
                    texte = f"{jour} {date}, {debut}–{fin}"
                else:
                    texte = f"{jour} {date}, {debut}–{fin}, salle {salle}"
                if matiere.dureePreparation:
                    passage = minutes_to_hhmm(horaire.debut + matiere.dureePreparation)
                    texte += f" (passage à {passage})"
                texte += f" : {matiere.nom}"

                eleves = ", ".join(
                    f"{etu.nom} ({etu.classe.nom}, G{etu.groupe})"
                    for _, etu in horaires_groupes[horaire]
                )

                render_activity_line(pdf, texte)
                render_eleve_line(pdf, f"Élèves : {eleves}")
                pdf.ln(1)

            pdf.ln(5)

        filepath = os.path.join(chemin_export, "colleurs", f"{colleur.nom}.pdf")
        pdf.output(filepath)



def export_resultats_html_json1(resultats, db, chemin_export):
    calendrier=db['calendrier']
    def get_heure(horaire):
        h = horaire.debut // 60
        m = horaire.debut % 60
        return f"{h:02}:{m:02}"

    DATA = {
        "classes": {str(c.id): {"nom": c.nom} for c in db["classes"].values()},
        "etudiants": {str(e.id): {"nom": e.nom, "classe": str(e.classe.id), "groupe": str(e.groupe)} for e in db["etudiants"].values()},
        "colleurs": {col.id: {"nom": col.nom ,"matieres": [m.id for m in col.matieres]} for col in db["colleurs"].values()},
        "matieres": {m.id: {"nom": m.nom, "classe": str(m.classe.id)} for m in db["matieres"].values()},
        "TPs": {tp.id: {"nom": tp.nom} for tp in db["TPs"].values()},
        "resultats_etudiants": defaultdict(lambda: defaultdict(list)),
        "resultats_colleurs": defaultdict(lambda: defaultdict(list)),
        "resultats_groupes": defaultdict(lambda: defaultdict(list)),
    }

    for semaine, planning in resultats.items():
        for etu, activites in planning.items():
            for a in activites:
                horaire = a[2]
                salle=horaire.salle
                date = calendrier.get_date(semaine, horaire.jour)
                jour_str = f"{horaire.jour} ({date.strftime('%d/%m')})"
                if isinstance(a[0], TP):
                    tp = a[0]
                    texte = f"{jour_str} {get_heure(horaire)} : {tp.nom}"
                    DATA["resultats_etudiants"][str(etu.id)][str(semaine)].append(texte)
                    DATA["resultats_groupes"][str(etu.groupe)][str(semaine)].append({
                        "etudiant": etu.nom,
                        "classe": db["classes"][etu.classe.id].nom,
                        "groupe": str(etu.groupe),
                        "matiere": tp.nom,
                        "horaire": f"{jour_str} {get_heure(horaire)}"
                    })
                else:
                    mat, col = a[0], a[1]
                    prep = ""
                    if mat.dureePreparation:
                        h = (horaire.debut - mat.dureePreparation) // 60
                        m = (horaire.debut - mat.dureePreparation) % 60
                        prep = f" (passage à {horaire.debut // 60:02}:{horaire.debut % 60:02})"
                        heure = f"{h:02}:{m:02}"
                    else:
                        heure = get_heure(horaire)
                    if salle=="":
                        texte = f"{jour_str} {heure} : {mat.nom} avec {col.nom}{prep}"
                    else:
                        texte = f"{jour_str} {heure} : {mat.nom} avec {col.nom}{prep} en salle {salle}"
                    DATA["resultats_etudiants"][str(etu.id)][str(semaine)].append(texte)
                    texte_colleur = f"{jour_str} {heure} : {etu.nom} ({db['classes'][etu.classe.id].nom}, groupe {etu.groupe})"
                    if salle:
                        texte_colleur += f" en salle {salle}"
                    DATA["resultats_colleurs"][str(col.id)][str(semaine)].append(texte_colleur)

                    DATA["resultats_groupes"][str(etu.groupe)][str(semaine)].append({
                        "etudiant": etu.nom,
                        "classe": str(etu.classe.id),
                        "groupe": str(etu.groupe),
                        "matiere": mat.nom,
                        "horaire": f"{jour_str} {heure}",
                        "colleur": col.id,
                        "salle": salle
                    })

    # Convertir les defaultdict en dict avec clés en chaînes
    for key in ["resultats_etudiants", "resultats_colleurs", "resultats_groupes"]:
        DATA[key] = {str(k): dict(v) for k, v in DATA[key].items()}
    os.makedirs("affichage_interactif", exist_ok=True)
    # Export
    with open(chemin_export+"/data.js", "w", encoding="utf-8") as f:
        f.write("const DATA = ")
        json.dump(DATA, f, ensure_ascii=False, indent=2)





def export_resultats_html_json(resultats, db, chemin_export):
    calendrier = db['calendrier']

    def get_heure(horaire):
        h = horaire.debut // 60
        m = horaire.debut % 60
        return f"{h:02}:{m:02}"

    DATA = {
        "classes": {str(c.id): {"nom": c.nom} for c in db["classes"].values()},
        "etudiants": {str(e.id): {"nom": e.nom, "classe": str(e.classe.id), "groupe": str(e.groupe)} for e in db["etudiants"].values()},
        "colleurs": {col.id: {"nom": col.nom ,"matieres": [m.id for m in col.matieres]} for col in db["colleurs"].values()},
        "matieres": {m.id: {"nom": m.nom, "classe": str(m.classe.id)} for m in db["matieres"].values()},
        "TPs": {tp.id: {"nom": tp.nom} for tp in db["TPs"].values()},
        "resultats_etudiants": defaultdict(lambda: defaultdict(list)),
        "resultats_colleurs": defaultdict(lambda: defaultdict(list)),
        "resultats_groupes": defaultdict(lambda: defaultdict(list)),
    }

    groupes_par_classe = defaultdict(set)

    for semaine, planning in resultats.items():
        for etu, activites in planning.items():
            groupes_par_classe[str(etu.classe.id)].add(str(etu.groupe))
            for a in activites:
                horaire = a[2]
                salle = horaire.salle
                date = calendrier.get_date(semaine, horaire.jour)
                jour_str = f"{horaire.jour} ({date.strftime('%d/%m')})"
                if isinstance(a[0], TP):
                    tp = a[0]
                    texte = f"{jour_str} {get_heure(horaire)} : {tp.nom}"
                    DATA["resultats_etudiants"][str(etu.id)][str(semaine)].append(texte)
                    DATA["resultats_groupes"][f"{etu.classe.id}-{etu.groupe}"][str(semaine)].append({
                        "etudiant": etu.nom,
                        "classe": db["classes"][etu.classe.id].nom,
                        "groupe": str(etu.groupe),
                        "matiere": tp.nom,
                        "horaire": f"{jour_str} {get_heure(horaire)}"
                    })
                else:
                    mat, col = a[0], a[1]
                    prep = ""
                    if mat.dureePreparation:
                        h = (horaire.debut - mat.dureePreparation) // 60
                        m = (horaire.debut - mat.dureePreparation) % 60
                        prep = f" (passage à {horaire.debut // 60:02}:{horaire.debut % 60:02})"
                        heure = f"{h:02}:{m:02}"
                    else:
                        heure = get_heure(horaire)
                    if salle == "":
                        texte = f"{jour_str} {heure} : {mat.nom} avec {col.nom}{prep}"
                    else:
                        texte = f"{jour_str} {heure} : {mat.nom} avec {col.nom}{prep} en salle {salle}"
                    DATA["resultats_etudiants"][str(etu.id)][str(semaine)].append(texte)
                    texte_colleur = f"{jour_str} {heure} : {etu.nom} ({db['classes'][etu.classe.id].nom}, groupe {etu.groupe})"
                    if salle:
                        texte_colleur += f" en salle {salle}"
                    DATA["resultats_colleurs"][str(col.id)][str(semaine)].append(texte_colleur)

                    DATA["resultats_groupes"][f"{etu.classe.id}-{etu.groupe}"][str(semaine)].append({
                        "etudiant": etu.nom,
                        "classe": str(etu.classe.id),
                        "groupe": str(etu.groupe),
                        "matiere": mat.nom,
                        "horaire": f"{jour_str} {heure}",
                        "colleur": col.id,
                        "salle": salle
                    })

    for key in ["resultats_etudiants", "resultats_colleurs", "resultats_groupes"]:
        DATA[key] = {str(k): dict(v) for k, v in DATA[key].items()}

    DATA["groupes_par_classe"] = {classe: sorted(list(groupes)) for classe, groupes in groupes_par_classe.items()}

    chemin_export.mkdir(parents=True, exist_ok=True)

    with open(os.path.join(chemin_export/"data.js"), "w", encoding="utf-8") as f:
        f.write("const DATA = ")
        json.dump(DATA, f, ensure_ascii=False, indent=2)

    with open(os.path.join(chemin_export/"Colloscope.html"), "w", encoding="utf-8") as f:
        f.write(r"""<!DOCTYPE html>
    <html lang="fr">
    <head>
    <meta charset="UTF-8">
    <title>Affichage des résultats</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        select, button { margin: 5px; padding: 5px; }
        #resultats { margin-top: 20px; white-space: pre-line; border: 1px solid #ccc; padding: 10px; }
    </style>
    </head>
    <body>
    <h2>Visualisation des résultats</h2>
    <label for="classeSelect">Classe :</label>
    <select id="classeSelect"></select>

    <label for="typeSelect">Type :</label>
    <select id="typeSelect">
        <option value="etudiant">Etudiant</option>
        <option value="colleur">Colleur</option>
        <option value="groupe">Groupe</option>
    </select>

    <label for="elementSelect">Élément :</label>
    <select id="elementSelect"></select>

    <label for="semaineSelect">Semaine :</label>
    <select id="semaineSelect"></select>

    <button onclick="afficherResultats()">Afficher</button>

    <div id="resultats"></div>

    <script src="data.js"></script>
    <script>
        const classeSelect = document.getElementById("classeSelect");
        const typeSelect = document.getElementById("typeSelect");
        const elementSelect = document.getElementById("elementSelect");
        const semaineSelect = document.getElementById("semaineSelect");
        const resultatsDiv = document.getElementById("resultats");

        function init() {
        classeSelect.innerHTML = "";
        for (const id in DATA.classes) {
            const opt = document.createElement("option");
            opt.value = id;
            opt.textContent = DATA.classes[id].nom;
            classeSelect.appendChild(opt);
        }
        updateElementSelect();
        }

        function updateElementSelect() {
        const type = typeSelect.value;
        const classeId = classeSelect.value;
        elementSelect.innerHTML = "";
        let elements = [];
        let noms = {};

        if (type === "etudiant") {
            for (const id in DATA.etudiants) {
            const e = DATA.etudiants[id];
            if (e.classe === classeId) {
                elements.push(id);
                noms[id] = e.nom;
            }
            }
        } else if (type === "colleur") {
            for (const id in DATA.colleurs) {
            const col = DATA.colleurs[id];
            for (const matId of col.matieres) {
                if (DATA.matieres[matId]?.classe === classeId) {
                elements.push(id);
                noms[id] = col.nom;
                break;
                }
            }
            }
        } else if (type === "groupe") {
            const groupes = (DATA.groupes_par_classe[classeId] || []).sort((a, b) => +a - +b);
            elements = groupes.map(g => `${classeId}-${g}`);
            for (const g of groupes) noms[`${classeId}-${g}`] = `Groupe ${g}`;
        }

        for (const id of elements) {
            const opt = document.createElement("option");
            opt.value = id;
            opt.textContent = noms[id];
            elementSelect.appendChild(opt);
        }
        updateSemaineSelect();
        }

        function updateSemaineSelect() {
        const type = typeSelect.value;
        const id = elementSelect.value;
        const data = DATA[`resultats_${type}s`]?.[id] || {};
        const semaines = Object.keys(data).sort((a, b) => +a - +b);

        semaineSelect.innerHTML = "";
        for (const s of semaines) {
            const opt = document.createElement("option");
            opt.value = s;
            opt.textContent = `Semaine ${s}`;
            semaineSelect.appendChild(opt);
        }
        }

        function afficherResultats() {
        const type = typeSelect.value;
        const id = elementSelect.value;
        const semaine = semaineSelect.value;
        const data = DATA[`resultats_${type}s`]?.[id]?.[semaine];

        if (!data) {
            resultatsDiv.textContent = "Aucune activité pour cette semaine.";
            return;
        }

        if (Array.isArray(data)) {
            if (typeof data[0] === 'string') {
            resultatsDiv.textContent = data.join("\n");
            } else {
            resultatsDiv.textContent = data.map(d => {
                return `- Étudiant : ${d.etudiant}\n  Matière : ${d.matiere}\n  Horaire : ${d.horaire}` +
                    (d.salle ? `\n  Salle : ${d.salle}` : '') +
                    (d.colleur ? `\n  Colleur : ${DATA.colleurs[d.colleur]?.nom || d.colleur}` : '');
            }).join("\n\n");
            }
        } else {
            resultatsDiv.textContent = JSON.stringify(data, null, 2);
        }
        }

        classeSelect.addEventListener("change", updateElementSelect);
        typeSelect.addEventListener("change", updateElementSelect);
        elementSelect.addEventListener("change", updateSemaineSelect);

        init();
    </script>
    </body>
    </html>
    """)



def export_resultats_groupes_xlsx(resultats, db, calendrier, chemin_export):
    wb = Workbook()
    for k,cl_id in enumerate(db['classes']):
        if k==0:
            ws = wb.active
            ws.title = f"Planning {db['classes'][cl_id].nom}"
        else:
            ws=wb.create_sheet(f"Planning {db['classes'][cl_id].nom}")


        row_cursor = 1



        # Regroupement par matière ID pour éviter les collisions
        matieres_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
        nom_matiere_par_id = {}
        toutes_les_semaines = set()

        for semaine, planning in resultats.items():
            toutes_les_semaines.add(semaine)
            for etu, activites in planning.items():
                if etu.classe.id==cl_id:
                    for act in activites:
                        obj, colleur, horaire = act
                        mat_id = obj.id
                        nom_matiere_par_id[mat_id] = obj.nom
                        clef_col = colleur.nom if colleur else ""
                        clef_hor = (horaire.jour, horaire.debut, horaire.duree, getattr(horaire, "salle", ""))
                        matieres_dict[mat_id][clef_col][clef_hor][semaine].append(etu)

            semaines = sorted(toutes_les_semaines)

        for mat_id, colleurs in matieres_dict.items():
            matiere = nom_matiere_par_id[mat_id]
            start_row = row_cursor

            # Ligne 1 : fusion et nom matière
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
            ws.cell(row_cursor, 1, matiere).font = Font(bold=True)
            ws.cell(row_cursor, 1).alignment = Alignment(horizontal="center")

            # Ligne 2 : en-têtes des colonnes
            ws.cell(row_cursor + 1, 1, "Colleur")
            ws.cell(row_cursor + 1, 2, "Horaire")
            ws.cell(row_cursor + 1, 3, "Salle")
            ws.cell(row_cursor + 1, 4, "Semaines")
            for i, semaine in enumerate(semaines):
                ws.cell(row_cursor + 1, 5 + i, f"S{semaine}").alignment = Alignment(horizontal="center")

            row_cursor += 2  # Décalage pour données

            for colleur, horaires in colleurs.items():
                for (jour, debut, duree, salle), semaines_dict in horaires.items():
                    heure_str = f"{jour} {debut//60:02}:{debut%60:02}"
                    ligne = [colleur, heure_str, salle, ""]

                    for semaine in semaines:
                        etudiants = semaines_dict.get(semaine, [])
                        if not etudiants:
                            ligne.append("")
                            continue

                        groupes = set(e.groupe for e in etudiants)
                        tous_du_meme_groupe = len(groupes) == 1
                        groupe = groupes.pop() if tous_du_meme_groupe else None

                        if tous_du_meme_groupe:
                            groupe_complet = [e for e in db["etudiants"].values()
                                            if e.groupe == groupe and e.classe == etudiants[0].classe]
                            if set(etudiants) == set(groupe_complet):
                                contenu = f"G{groupe}"
                            else:
                                contenu = "\n".join(sorted(e.nom for e in etudiants))
                        else:
                            contenu = "\n".join(sorted(e.nom for e in etudiants))

                        ligne.append(contenu)

                    ws.append(ligne)
                    row_cursor += 1

            row_cursor += 2  # Espace après chaque matière

        # Feuille secondaire : liste des groupes
        ws2 = wb.create_sheet(f"Groupes {db['classes'][cl_id].nom}")
        ws2.append(["Nom", "Groupe"])

        for etu in sorted(db["etudiants"].values(), key=lambda e: (e.classe, e.groupe, e.nom)):
            if etu.classe.id==cl_id:
                ws2.append([ etu.nom, etu.groupe])

    dirpath = os.path.dirname(chemin_export)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)

    wb.save(chemin_export)
    return chemin_export



if __name__=="__main__":

    db=load_all("PCSI_Test.xml")
    resultats=load_solution_from_xml("Test_solution.xml",db)

    #export_resultats_pdf("Resultats_pdf",resultats,db)
    export_resultats_html_json(resultats, db, "affichage_interactif")

    #export_resultats_groupes_xlsx(resultats,db,db['calendrier'],"tableau_extension.xlsx")
